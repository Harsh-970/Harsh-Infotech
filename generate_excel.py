"""
Harsh Infotech Website Data Extractor
Generates website-data.xlsx with all content organized by page/section.
Run: python generate_excel.py
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print("openpyxl found, proceeding...")
except ImportError:
    import subprocess, sys
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Color palette ───────────────────────────────────────────
GOLD       = "D4AF37"
BLACK      = "0A0A0A"
DARK_GRAY  = "1A1A1A"
MID_GRAY   = "2A2A2A"
LIGHT_GRAY = "F5F5F5"
WHITE      = "FFFFFF"
HEADER_BG  = "1C1C1C"

def make_header_style(bold=True, bg=HEADER_BG, fg=GOLD, size=11) -> dict:
    return dict(
        font=Font(name="Calibri", bold=bold, color=fg, size=size),
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=Border(
            bottom=Side(style="medium", color=GOLD),
            right=Side(style="thin", color="3A3A3A"),
        )
    )

def make_row_style(alt=False) -> dict:
    bg = "1E1E1E" if alt else "161616"
    return dict(
        font=Font(name="Calibri", color=WHITE, size=10),
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        border=Border(
            bottom=Side(style="thin", color="2A2A2A"),
            right=Side(style="thin", color="2A2A2A"),
        )
    )

def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def write_sheet(ws, headers: list, rows: list, col_widths: list = None):
    """Write headers + rows to a worksheet with styling."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    # Write headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        apply_style(cell, make_header_style())
    ws.row_dimensions[1].height = 22

    # Write data rows
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            apply_style(cell, make_row_style(alt=ri % 2 == 0))
        ws.row_dimensions[ri].height = 40

    # Column widths
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    else:
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 30

    # Freeze header row
    ws.freeze_panes = "A2"

# ── Build workbook ──────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════
# 1. HOME PAGE
# ════════════════════════════════════════════
ws = wb.create_sheet("Home_Page")
headers = ["Section", "Field", "Content", "Notes"]
rows = [
    ["Hero", "Main Heading (Line 1)", "Reliably Great", "Bold, white text"],
    ["Hero", "Main Heading (Line 2)", "Efficiently Fast", "Bold, white/40 (muted)"],
    ["Hero", "Subheading", "Complete Tally Solutions for Modern Businesses", "Strong/bold, white"],
    ["Hero", "Description", "We help businesses simplify accounting, automate workflows, and scale operations with powerful Tally solutions.", "white/60"],
    ["Hero", "CTA Button", "Get Started", "White button, black text"],
    ["Hero", "Trust Line", "Trusted by 100+ businesses across India", "Small, muted text below button"],
    ["Hero", "Right Side", "Logo (large, dominant)", "~460–520px on desktop, gold glow"],
    ["", "", "", ""],
    ["Services Section", "Section Heading", "Our Services", "Center aligned"],
    ["Services Section", "Section Description", "We provide complete Tally solutions to help businesses streamline operations and improve efficiency.", "Center, muted"],
    ["Services Section", "Card 1 – Title", "Tally Prime License", "Links to /services.html#tally-license"],
    ["Services Section", "Card 1 – Description", "Get genuine Tally licenses (Single & Multi User) with complete setup and guidance.", ""],
    ["Services Section", "Card 2 – Title", "Tally on Cloud", "Links to /services.html#tally-cloud"],
    ["Services Section", "Card 2 – Description", "Access your business data anytime, anywhere with secure and reliable cloud solutions.", "Has 'Start Free Trial' CTA"],
    ["Services Section", "Card 2 – CTA", "Start Free Trial", "Gold text, scrolls to contact"],
    ["Services Section", "Card 3 – Title", "Tally Customization", "Links to /services.html#tally-customization"],
    ["Services Section", "Card 3 – Description", "Customize Tally according to your business workflow using advanced TDL solutions.", ""],
    ["", "", "", ""],
    ["More Solutions Section", "Section Heading", "More Solutions", ""],
    ["More Solutions Section", "Section Description", "Everything you need to run your business infrastructure smoothly.", ""],
    ["More Solutions Section", "Card 1 – Title", "VPS Hosting", ""],
    ["More Solutions Section", "Card 1 – Description", "High-performance virtual private servers for your business applications.", ""],
    ["More Solutions Section", "Card 2 – Title", "Tally AMC", ""],
    ["More Solutions Section", "Card 2 – Description", "Annual maintenance contracts to keep your Tally running smoothly.", ""],
    ["More Solutions Section", "Card 3 – Title", "Excel Integration", ""],
    ["More Solutions Section", "Card 3 – Description", "Seamlessly import your Excel data directly into Tally.", ""],
    ["More Solutions Section", "Card 4 – Title", "Data Migration", ""],
    ["More Solutions Section", "Card 4 – Description", "Secure and reliable transfer of your business data.", ""],
    ["More Solutions Section", "Card 5 – Title", "Hardware Support", ""],
    ["More Solutions Section", "Card 5 – Description", "Complete IT infrastructure support for your business.", ""],
    ["", "", "", ""],
    ["Products Section", "Section Heading", "Our Products", ""],
    ["Products Section", "Section Description", "Premium hardware and IT infrastructure for ambitious businesses.", ""],
    ["Products Section", "Card 1 – Title", "Enterprise Servers", "Links to /products.html#servers"],
    ["Products Section", "Card 1 – Description", "High-performance server solutions for demanding business workloads.", ""],
    ["Products Section", "Card 2 – Title", "Professional Workstations", "Links to /products.html#workstations"],
    ["Products Section", "Card 2 – Description", "Precision-engineered workstations for maximum productivity.", ""],
    ["Products Section", "Card 3 – Title", "Commercial Printers", "Links to /products.html#printers"],
    ["Products Section", "Card 3 – Description", "Industrial-grade printing solutions for high-volume needs.", ""],
    ["Products Section", "Card 4 – Title", "High-Speed Scanners", "Links to /products.html#scanners"],
    ["Products Section", "Card 4 – Description", "Ultra-fast document scanning for efficient workflows.", ""],
    ["", "", "", ""],
    ["Why Choose Us", "Section Heading", "Why Businesses Choose Harsh Infotech", "Harsh Infotech in gold"],
    ["Why Choose Us", "Description", "We deliver reliable Tally solutions designed to simplify operations, improve accuracy, and scale your business efficiently.", ""],
    ["Why Choose Us", "Feature 1 – Title", "Tally Expertise", "Gold icon: Database"],
    ["Why Choose Us", "Feature 1 – Description", "Complete support for Tally Prime, customization, and integrations tailored to your business workflow.", ""],
    ["Why Choose Us", "Feature 2 – Title", "Fast Implementation", "Gold icon: Zap"],
    ["Why Choose Us", "Feature 2 – Description", "Quick setup and deployment with minimal downtime so your business runs smoothly.", ""],
    ["Why Choose Us", "Feature 3 – Title", "Reliable Support", "Gold icon: Shield"],
    ["Why Choose Us", "Feature 3 – Description", "Ongoing AMC support with quick issue resolution and expert guidance.", ""],
    ["Why Choose Us", "Floating Stat", "100+ Businesses Served", "Glass card overlapping bottom-right of image"],
    ["", "", "", ""],
    ["Testimonials Section", "Section Heading", "What Our Clients Say", ""],
    ["Testimonials Section", "Section Description", "Trusted by 100+ businesses across India for Tally solutions, cloud services, and IT support.", ""],
]
write_sheet(ws, headers, rows, [22, 30, 60, 30])

# ════════════════════════════════════════════
# 2. SERVICES_MAIN
# ════════════════════════════════════════════
ws2 = wb.create_sheet("Services_Main")
headers2 = ["Service", "Sub-Section", "Field", "Content", "Price / CTA", "WhatsApp Link"]
rows2 = [
    ["Page Header", "–", "Page Title", "Tally Services", "–", "–"],
    ["Page Header", "–", "Page Description", "Genuine Tally licenses, cloud access, and custom solutions — everything your business needs in one place.", "–", "–"],
    ["", "", "", "", "", ""],
    ["Tally Prime License", "Section Subtitle", "–", "Genuine licensing with complete setup and support", "–", "services.html#tally-license"],
    ["Tally Prime License", "Description", "–", "Get genuine Tally Prime licenses (Single & Multi User) directly from Tally Solutions certified partners. Full setup, activation, migration and training included. Annual renewal pricing is the same as license cost.", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Plan Name", "Single User (Silver)", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Target Audience", "Best for individual accountants and small businesses with one Tally user.", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Monthly Price", "Monthly Subscription", "₹750 / mo", "–"],
    ["Tally Prime License", "Silver Plan", "Yearly Price", "Yearly Subscription (Save 2 months)", "₹7,500 / yr", "–"],
    ["Tally Prime License", "Silver Plan", "Perpetual Price", "Perpetual License (One time)", "₹22,500", "–"],
    ["Tally Prime License", "Silver Plan", "Renewal Price", "Annual Renewal (Perpetual)", "₹22,500 / yr", "–"],
    ["Tally Prime License", "Silver Plan", "Feature 1", "GST Ready Billing", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Feature 2", "Inventory Management", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Feature 3", "Bank Reconciliation", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Feature 4", "1 Concurrent User", "–", "–"],
    ["Tally Prime License", "Silver Plan", "Feature 5", "Free Setup Assistance", "–", "–"],
    ["Tally Prime License", "Silver Plan", "CTA Button", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in Tally Prime Silver (Single User) License"],
    ["", "", "", "", "", ""],
    ["Tally Prime License", "Gold Plan", "Plan Name", "Multi User (Gold)", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Target Audience", "Ideal for businesses with multiple users needing simultaneous Tally access.", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Monthly Price", "Monthly Subscription", "₹1,500 / mo", "–"],
    ["Tally Prime License", "Gold Plan", "Yearly Price", "Yearly Subscription (Save 2 months)", "₹15,000 / yr", "–"],
    ["Tally Prime License", "Gold Plan", "Perpetual Price", "Perpetual License (One time)", "₹67,500", "–"],
    ["Tally Prime License", "Gold Plan", "Renewal Price", "Annual Renewal (Perpetual)", "₹67,500 / yr", "–"],
    ["Tally Prime License", "Gold Plan", "Feature 1", "Everything in Silver", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Feature 2", "Unlimited Concurrent Users", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Feature 3", "Multi-Company Support", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Feature 4", "Remote Access Ready", "–", "–"],
    ["Tally Prime License", "Gold Plan", "Feature 5", "Priority Support", "–", "–"],
    ["Tally Prime License", "Gold Plan", "CTA Button", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in Tally Prime Gold (Multi User) License"],
    ["Tally Prime License", "Renewal Note", "–", "Annual renewal cost for perpetual licenses is identical to the original license price. Subscription plans renew at their respective monthly/yearly rates with no hidden charges.", "–", "–"],
    ["", "", "", "", "", ""],
    ["Tally on Cloud", "Section Subtitle", "–", "Access your business data from anywhere, anytime", "–", "services.html#tally-cloud"],
    ["Tally on Cloud", "Description", "–", "Run Tally Prime on a secure, always-online VPS — no local install headaches. Access from any device, any location, with your data backed up and protected 24/7.", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "Users", "1–2 Users", "₹1,999 / mo", "–"],
    ["Tally on Cloud", "Starter Plan", "Feature 1", "2 vCPU / 4 GB RAM", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "Feature 2", "50 GB SSD Storage", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "Feature 3", "Remote Desktop Access", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "Feature 4", "Daily Auto Backup", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "Feature 5", "99.9% Uptime SLA", "–", "–"],
    ["Tally on Cloud", "Starter Plan", "CTA", "Get Started", "–", "wa.me/917558604483?text=Hi, I am interested in Tally on Cloud - Starter plan"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Users", "Up to 5 Users", "₹3,499 / mo", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Feature 1", "4 vCPU / 8 GB RAM", "–", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Feature 2", "100 GB SSD Storage", "–", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Feature 3", "Remote Desktop Access", "–", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Feature 4", "Daily + Weekly Backup", "–", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "Feature 5", "Priority Support", "–", "–"],
    ["Tally on Cloud", "Business Plan (Most Popular)", "CTA", "Get Started", "–", "wa.me/917558604483?text=Hi, I am interested in Tally on Cloud - Business plan"],
    ["Tally on Cloud", "Enterprise Plan", "Users", "10+ Users", "Custom", "–"],
    ["Tally on Cloud", "Enterprise Plan", "Feature 1", "Dedicated Server", "–", "–"],
    ["Tally on Cloud", "Enterprise Plan", "Feature 2", "Custom Storage", "–", "–"],
    ["Tally on Cloud", "Enterprise Plan", "Feature 3", "Dedicated Support Line", "–", "–"],
    ["Tally on Cloud", "Enterprise Plan", "Feature 4", "Hourly Backups", "–", "–"],
    ["Tally on Cloud", "Enterprise Plan", "Feature 5", "SLA Guarantee", "–", "–"],
    ["Tally on Cloud", "Enterprise Plan", "CTA", "Get Started", "–", "wa.me/917558604483?text=Hi, I am interested in Tally on Cloud - Enterprise plan"],
    ["Tally on Cloud", "Inclusion Note", "–", "All plans include Tally Prime software access, remote desktop connectivity, firewall protection, and technical setup by our team. Your own Tally license is required or can be bundled.", "–", "–"],
    ["", "", "", "", "", ""],
    ["Tally Customization", "Section Subtitle", "–", "TDL-powered workflows tailored to your business", "–", "services.html#tally-customization"],
    ["Tally Customization", "Description", "–", "We build custom Tally Definition Language (TDL) modules that extend Tally Prime to match your exact business workflow — from custom reports and invoices to automated workflows and integrations.", "–", "–"],
    ["Tally Customization", "Module 1", "Title", "Custom Invoice & Print Formats", "–", "–"],
    ["Tally Customization", "Module 1", "Description", "Design invoices, delivery challans, and receipts exactly matching your brand and format requirements.", "–", "–"],
    ["Tally Customization", "Module 1", "Features", "Logo & Brand Integration | GST Compliant Formats | Multi-language Support | Auto Calculation Fields", "–", "–"],
    ["Tally Customization", "Module 2", "Title", "Custom Reports & MIS", "–", "–"],
    ["Tally Customization", "Module 2", "Description", "Build management reports, stock reports, ledger summaries, and dashboards tailored to your decision-making needs.", "–", "–"],
    ["Tally Customization", "Module 2", "Features", "Profit & Loss Variants | Stock Ageing Analysis | Custom Ledger Reports | Excel Export Ready", "–", "–"],
    ["Tally Customization", "Module 3", "Title", "Workflow Automation", "–", "–"],
    ["Tally Customization", "Module 3", "Description", "Automate repetitive tasks — voucher approvals, party alerts, payment reminders, and more.", "–", "–"],
    ["Tally Customization", "Module 3", "Features", "Auto Voucher Posting | Email / SMS Alerts | Approval Workflows | Scheduled Jobs", "–", "–"],
    ["Tally Customization", "Module 4", "Title", "Third-Party Integration", "–", "–"],
    ["Tally Customization", "Module 4", "Description", "Connect Tally with eCommerce platforms, HRMS, CRM, or any external system via API or import tools.", "–", "–"],
    ["Tally Customization", "Module 4", "Features", "GST Portal Filing | eCommerce Sync | Payroll Integration | Bank Statement Import", "–", "–"],
    ["Tally Customization", "CTA Panel", "Heading", "Custom Pricing — Built for Your Scope", "Custom", "–"],
    ["Tally Customization", "CTA Panel", "Description", "Every customization is unique. Share your requirements and we'll quote based on complexity, modules, and timeline.", "–", "–"],
    ["Tally Customization", "CTA Panel", "Button", "Discuss Your Project", "–", "wa.me/917558604483?text=Hi, I need a custom Tally TDL solution for my business"],
]
write_sheet(ws2, headers2, rows2, [25, 28, 22, 55, 18, 45])

# ════════════════════════════════════════════
# 3. MORE_SERVICES
# ════════════════════════════════════════════
ws3 = wb.create_sheet("More_Services")
headers3 = ["Service", "Field", "Content", "Price / CTA", "WhatsApp Link"]
rows3 = [
    ["Page Header", "Page Title", "Digital Services", "–", "more-services.html"],
    ["Page Header", "Page Description", "Explore our comprehensive suite of advanced technology solutions tailored for scaling your enterprise.", "–", "–"],
    ["", "", "", "", ""],
    ["VPS Hosting", "Title", "VPS (Virtual Private Server)", "₹1,999 / mo", "more-services.html#vps"],
    ["VPS Hosting", "Description", "High-performance cloud servers for running business applications securely and efficiently.", "–", "–"],
    ["VPS Hosting", "Feature 1", "99.9% Uptime Guarantee", "–", "–"],
    ["VPS Hosting", "Feature 2", "SSD Backed Storage", "–", "–"],
    ["VPS Hosting", "Feature 3", "24/7 Monitoring", "–", "–"],
    ["VPS Hosting", "Feature 4", "Full Root Access", "–", "–"],
    ["VPS Hosting", "CTA Button", "View Plans", "–", "wa.me/917558604483?text=Hi, I am interested in VPS (Virtual Private Server)"],
    ["", "", "", "", ""],
    ["Tally AMC & Support", "Title", "Tally AMC & Support", "₹4,500 / yr", "more-services.html#amc"],
    ["Tally AMC & Support", "Description", "Reliable annual maintenance and support for smooth business operations.", "–", "–"],
    ["Tally AMC & Support", "Feature 1", "Priority Remote Support", "–", "–"],
    ["Tally AMC & Support", "Feature 2", "Data Corruption Fixes", "–", "–"],
    ["Tally AMC & Support", "Feature 3", "Version Upgrades", "–", "–"],
    ["Tally AMC & Support", "Feature 4", "On-site Visits", "–", "–"],
    ["Tally AMC & Support", "CTA Button", "Get Contract", "–", "wa.me/917558604483?text=Hi, I am interested in Tally AMC & Support"],
    ["", "", "", "", ""],
    ["Excel to Tally Integration", "Title", "Excel to Tally Integration", "Custom", "more-services.html#excel"],
    ["Excel to Tally Integration", "Description", "Seamlessly import and manage your vast Excel data directly into Tally.", "–", "–"],
    ["Excel to Tally Integration", "Feature 1", "Bulk Ledger Creation", "–", "–"],
    ["Excel to Tally Integration", "Feature 2", "Voucher Imports", "–", "–"],
    ["Excel to Tally Integration", "Feature 3", "Custom Mapping", "–", "–"],
    ["Excel to Tally Integration", "Feature 4", "Error Validation", "–", "–"],
    ["Excel to Tally Integration", "CTA Button", "Request Demo", "–", "wa.me/917558604483?text=Hi, I am interested in Excel to Tally Integration"],
    ["", "", "", "", ""],
    ["Data Migration & Setup", "Title", "Data Migration & Setup", "Custom", "more-services.html#data"],
    ["Data Migration & Setup", "Description", "Secure transfer and setup of your existing business data to new systems or cloud.", "–", "–"],
    ["Data Migration & Setup", "Feature 1", "Zero Data Loss guarantee", "–", "–"],
    ["Data Migration & Setup", "Feature 2", "Downtime Minimization", "–", "–"],
    ["Data Migration & Setup", "Feature 3", "Secure Encryption", "–", "–"],
    ["Data Migration & Setup", "Feature 4", "Post-Migration Audit", "–", "–"],
    ["Data Migration & Setup", "CTA Button", "Start Migration", "–", "wa.me/917558604483?text=Hi, I am interested in Data Migration & Setup"],
    ["", "", "", "", ""],
    ["Hardware & System Support", "Title", "Hardware & System Support", "Hourly", "more-services.html#hardware"],
    ["Hardware & System Support", "Description", "Complete assistance for business systems and physical IT infrastructure.", "–", "–"],
    ["Hardware & System Support", "Feature 1", "Network Troubleshooting", "–", "–"],
    ["Hardware & System Support", "Feature 2", "Hardware Repair", "–", "–"],
    ["Hardware & System Support", "Feature 3", "System Upgrades", "–", "–"],
    ["Hardware & System Support", "Feature 4", "Peripheral Setup", "–", "–"],
    ["Hardware & System Support", "CTA Button", "Contact Support", "–", "wa.me/917558604483?text=Hi, I am interested in Hardware & System Support"],
]
write_sheet(ws3, headers3, rows3, [28, 22, 55, 15, 45])

# ════════════════════════════════════════════
# 4. PRODUCTS
# ════════════════════════════════════════════
ws4 = wb.create_sheet("Products")
headers4 = ["Product", "Field", "Content", "Price / CTA", "Page Link"]
rows4 = [
    ["Page Header", "Page Title", "Hardware Products", "–", "products.html"],
    ["Page Header", "Page Description", "Premium IT hardware and infrastructure solutions for modern businesses.", "–", "–"],
    ["", "", "", "", ""],
    ["Enterprise Servers", "Title", "Enterprise Servers", "Starting at ₹85,000", "products.html#servers"],
    ["Enterprise Servers", "Description", "High-performance server solutions for demanding business workloads.", "–", "–"],
    ["Enterprise Servers", "Feature 1", "Dual Xeon Processors", "–", "–"],
    ["Enterprise Servers", "Feature 2", "ECC Memory Support", "–", "–"],
    ["Enterprise Servers", "Feature 3", "RAID Storage Arrays", "–", "–"],
    ["Enterprise Servers", "Feature 4", "Remote Management", "–", "–"],
    ["Enterprise Servers", "CTA", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in Enterprise Servers"],
    ["", "", "", "", ""],
    ["Professional Workstations", "Title", "Professional Workstations", "Starting at ₹45,000", "products.html#workstations"],
    ["Professional Workstations", "Description", "Precision-engineered workstations for maximum productivity.", "–", "–"],
    ["Professional Workstations", "Feature 1", "Latest Gen Processors", "–", "–"],
    ["Professional Workstations", "Feature 2", "Professional GPU Options", "–", "–"],
    ["Professional Workstations", "Feature 3", "NVMe SSD Storage", "–", "–"],
    ["Professional Workstations", "Feature 4", "Multi-Monitor Support", "–", "–"],
    ["Professional Workstations", "CTA", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in Professional Workstations"],
    ["", "", "", "", ""],
    ["Commercial Printers", "Title", "Commercial Printers", "Starting at ₹12,000", "products.html#printers"],
    ["Commercial Printers", "Description", "Industrial-grade printing solutions for high-volume needs.", "–", "–"],
    ["Commercial Printers", "Feature 1", "50+ PPM Print Speed", "–", "–"],
    ["Commercial Printers", "Feature 2", "Network Ready", "–", "–"],
    ["Commercial Printers", "Feature 3", "Duplex Printing", "–", "–"],
    ["Commercial Printers", "Feature 4", "High Capacity Trays", "–", "–"],
    ["Commercial Printers", "CTA", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in Commercial Printers"],
    ["", "", "", "", ""],
    ["High-Speed Scanners", "Title", "High-Speed Scanners", "Starting at ₹8,500", "products.html#scanners"],
    ["High-Speed Scanners", "Description", "Ultra-fast document scanning for efficient workflows.", "–", "–"],
    ["High-Speed Scanners", "Feature 1", "ADF (Auto Doc Feeder)", "–", "–"],
    ["High-Speed Scanners", "Feature 2", "OCR Ready", "–", "–"],
    ["High-Speed Scanners", "Feature 3", "Cloud Integration", "–", "–"],
    ["High-Speed Scanners", "Feature 4", "Double-sided Scanning", "–", "–"],
    ["High-Speed Scanners", "CTA", "Get Quote", "–", "wa.me/917558604483?text=Hi, I am interested in High-Speed Scanners"],
]
write_sheet(ws4, headers4, rows4, [28, 22, 55, 20, 30])

# ════════════════════════════════════════════
# 5. ABOUT PAGE
# ════════════════════════════════════════════
ws5 = wb.create_sheet("About_Page")
headers5 = ["Section", "Field", "Content", "Notes"]
rows5 = [
    ["Page Header", "Page Title", "About Harsh Infotech", "–"],
    ["Page Header", "Page Description", "Dedicated to building secure, reliable, and powerful IT infrastructure for enterprises across India.", "–"],
    ["", "", "", ""],
    ["About Section", "Heading", "We Are Harsh Infotech", "–"],
    ["About Section", "Description", "Founded with the mission to make enterprise-grade technology accessible to all businesses, Harsh Infotech delivers reliable Tally solutions, IT infrastructure, and cloud services across India.", "–"],
    ["", "", "", ""],
    ["What We Do", "Section Heading", "What We Do", "3-column grid"],
    ["What We Do", "Card 1 – Title", "Tally Solutions", "–"],
    ["What We Do", "Card 1 – Description", "Complete Tally Prime licensing, customization, and cloud deployment.", "–"],
    ["What We Do", "Card 2 – Title", "IT Infrastructure", "–"],
    ["What We Do", "Card 2 – Description", "Servers, workstations, printers, and scanners for growing businesses.", "–"],
    ["What We Do", "Card 3 – Title", "Business Support", "–"],
    ["What We Do", "Card 3 – Description", "AMC contracts, data migration, and ongoing technical assistance.", "–"],
    ["", "", "", ""],
    ["Team Section", "Section Heading", "Our Team", "–"],
    ["Team Section", "Member 1 – Name", "Harsh Shah", "–"],
    ["Team Section", "Member 1 – Role", "Founder & CEO", "–"],
    ["Team Section", "Member 1 – Description", "Certified Tally partner and IT infrastructure specialist with years of industry experience.", "–"],
    ["Team Section", "Member 2 – Name", "Team Member", "–"],
    ["Team Section", "Member 2 – Role", "Technical Lead", "–"],
    ["Team Section", "Member 2 – Description", "Expert in Tally customization and cloud deployments.", "–"],
    ["Team Section", "Member 3 – Name", "Team Member", "–"],
    ["Team Section", "Member 3 – Role", "Support Engineer", "–"],
    ["Team Section", "Member 3 – Description", "Handles AMC, data migration, and on-site support.", "–"],
    ["", "", "", ""],
    ["Contact Section", "Section Heading", "Reach Out", "–"],
    ["Contact Section", "Email Label", "Email", "–"],
    ["Contact Section", "Email Value", "harshinfotech2005@gmail.com", "–"],
    ["Contact Section", "Phone Label", "Phone", "–"],
    ["Contact Section", "Phone Value 1", "7558604483", "–"],
    ["Contact Section", "Phone Value 2", "8828275219", "–"],
    ["Contact Section", "Location Label", "Location", "–"],
    ["Contact Section", "Location Value", "Mumbai, India", "–"],
]
write_sheet(ws5, headers5, rows5, [22, 28, 55, 25])

# ════════════════════════════════════════════
# 6. TESTIMONIALS
# ════════════════════════════════════════════
ws6 = wb.create_sheet("Testimonials")
headers6 = ["#", "Name", "Role / Title", "Testimonial Text", "Initials", "Notes"]
rows6 = [
    ["Section Header", "–", "–", "What Our Clients Say", "–", "Section heading"],
    ["Section Subtitle", "–", "–", "Trusted by 100+ businesses across India for Tally solutions, cloud services, and IT support.", "–", "Muted subheading"],
    ["", "", "", "", "", ""],
    [1, "Rajesh Sharma", "Business Owner", "Harsh Infotech helped us streamline our Tally operations and saved hours of manual work.", "RS", "Column 1"],
    [2, "Priya Mehta", "Accountant", "Excellent support and fast implementation. Their Tally customization is top-notch.", "PM", "Column 1"],
    [3, "Amit Verma", "Operations Manager", "Smooth migration and zero data loss. Highly recommended for growing companies.", "AV", "Column 1 / Column 3"],
    [4, "Sneha Patil", "Finance Head", "Reliable AMC support and quick issue resolution every time.", "SP", "Column 2"],
    [5, "Kunal Shah", "Startup Founder", "Their cloud Tally setup allowed us to work from anywhere seamlessly.", "KS", "Column 2"],
    [6, "Neha Gupta", "Admin Manager", "Professional team with deep knowledge of business workflows.", "NG", "Column 2 / Column 3"],
]
write_sheet(ws6, headers6, rows6, [12, 22, 22, 60, 10, 20])

# ════════════════════════════════════════════
# 7. CONTACT INFO
# ════════════════════════════════════════════
ws7 = wb.create_sheet("Contact_Info")
headers7 = ["Type", "Label", "Value", "Used In", "Notes"]
rows7 = [
    ["Email", "Primary Email", "harshinfotech2005@gmail.com", "Footer, About page", "Main contact email"],
    ["Phone", "Primary Phone", "7558604483", "Footer, About page, WhatsApp links", "Also used as WhatsApp number"],
    ["Phone", "Secondary Phone", "8828275219", "Footer, About page", "–"],
    ["Location", "City", "Mumbai", "Footer, About page", "–"],
    ["Location", "Country", "India", "Footer, About page", "–"],
    ["WhatsApp", "WhatsApp Base URL", "https://wa.me/917558604483", "All CTA buttons", "Prefix for all WhatsApp links"],
    ["Social", "Twitter/X", "–", "Footer", "Icon present, link not set"],
    ["Social", "GitHub", "–", "Footer", "Icon present, link not set"],
    ["Social", "LinkedIn", "–", "Footer", "Icon present, link not set"],
    ["", "", "", "", ""],
    ["Footer", "Brand Name", "Harsh Infotech Consultancy Services", "Footer", "Full legal name"],
    ["Footer", "Tagline", "Building the future of digital infrastructure. Premium solutions for ambitious companies.", "Footer", "–"],
    ["Footer", "Newsletter Label", "Subscribe to our latest updates.", "Footer", "Newsletter section"],
    ["Footer", "Newsletter Placeholder", "Email address", "Footer", "Input placeholder"],
    ["Footer", "Newsletter CTA", "Join", "Footer", "Button text"],
    ["Footer", "Copyright", "© 2024 Harsh Infotech. All rights reserved.", "Footer", "–"],
    ["Footer", "Quick Links", "Home | Services | About Us | Careers", "Footer", "Quick links column"],
]
write_sheet(ws7, headers7, rows7, [15, 25, 40, 25, 25])

# ════════════════════════════════════════════
# 8. NAVIGATION
# ════════════════════════════════════════════
ws8 = wb.create_sheet("Navigation")
headers8 = ["Menu Item", "Label", "URL / Link", "Page", "Type", "Notes"]
rows8 = [
    ["Navbar Brand", "Harsh Infotech Consultancy Services", "/", "All pages", "Logo + Text", "Logo image + brand name"],
    ["", "", "", "", "", ""],
    ["Main Nav", "Home", "/index.html", "All pages", "Link", "–"],
    ["Main Nav", "Services", "Dropdown", "All pages", "Dropdown", "Opens dropdown menu"],
    ["Main Nav → Services Dropdown", "Tally Prime License", "/services.html#tally-license", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Services Dropdown", "Tally on Cloud", "/services.html#tally-cloud", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Services Dropdown", "Tally Customization", "/services.html#tally-customization", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Services Dropdown", "More Services", "/more-services.html", "All pages", "Dropdown Item", "Divider before this item"],
    ["Main Nav", "Products", "Dropdown", "All pages", "Dropdown", "Opens dropdown menu"],
    ["Main Nav → Products Dropdown", "Enterprise Servers", "/products.html#servers", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Products Dropdown", "Professional Workstations", "/products.html#workstations", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Products Dropdown", "Commercial Printers", "/products.html#printers", "All pages", "Dropdown Item", "–"],
    ["Main Nav → Products Dropdown", "High-Speed Scanners", "/products.html#scanners", "All pages", "Dropdown Item", "–"],
    ["Main Nav", "About", "/about.html", "All pages", "Link", "–"],
    ["Main Nav", "Contact", "/#contact", "All pages", "Link", "Scrolls to contact section on homepage"],
    ["", "", "", "", "", ""],
    ["Top-Right CTA (Dynamic)", "Our Services", "/#services", "Homepage – 'home' section", "CTA Button", "Shows when at top of home"],
    ["Top-Right CTA (Dynamic)", "Our Products", "/#products", "Homepage – 'services' section", "CTA Button", "Changes as user scrolls"],
    ["Top-Right CTA (Dynamic)", "About", "/#about", "Homepage – 'products' section", "CTA Button", "–"],
    ["Top-Right CTA (Dynamic)", "Contact", "/#contact", "Homepage – 'about' section", "CTA Button", "–"],
    ["Top-Right CTA (Dynamic)", "More Services", "/more-services.html", "services.html", "CTA Button", "–"],
    ["Top-Right CTA (Dynamic)", "About Us", "/about.html", "Other pages (default)", "CTA Button", "–"],
    ["", "", "", "", "", ""],
    ["Page Routing", "Home", "/index.html → src/main.tsx → App.tsx", "–", "Page", "–"],
    ["Page Routing", "Services (Main)", "/services.html → src/main-services.tsx → MainServicesApp.tsx", "–", "Page", "3 core Tally services"],
    ["Page Routing", "More Services", "/more-services.html → src/more-services.tsx → ServicesApp.tsx", "–", "Page", "VPS, AMC, etc."],
    ["Page Routing", "Products", "/products.html → src/products.tsx → ProductsApp.tsx", "–", "Page", "–"],
    ["Page Routing", "About", "/about.html → src/about.tsx → AboutApp.tsx", "–", "Page", "–"],
    ["", "", "", "", "", ""],
    ["Mobile Nav", "Hamburger Menu", "Toggles mobile menu", "All pages mobile", "Button", "Shows on mobile/tablet only"],
    ["Mobile Nav", "All main nav items", "Same as desktop", "Mobile menu", "Links", "Stacked vertically in full-screen overlay"],
]
write_sheet(ws8, headers8, rows8, [30, 35, 45, 22, 15, 28])

# ── Save file ────────────────────────────────────────────────
output_path = r"d:\Harsh Documents\Projects\Webiste\google\harsh-infotech\website-data.xlsx"
wb.save(output_path)
print(f"\n✅ Excel file created successfully!")
print(f"   Location: {output_path}")
print(f"   Sheets: {', '.join(wb.sheetnames)}")
